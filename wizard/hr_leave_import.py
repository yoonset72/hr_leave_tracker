import base64
import csv
import io
import logging
from datetime import date

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError

_logger = logging.getLogger(__name__)

# Optional Excel libraries
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class HrLeaveImport(models.TransientModel):
    _name = 'hr.leave.import'
    _description = 'Import Historical Leave Data'

    import_file = fields.Binary(
        string='CSV or Excel File',
        required=True,
        help="Upload CSV or Excel file with historical leave data"
    )
    import_filename = fields.Char(string='Filename')
    file_type = fields.Selection([
        ('csv', 'CSV File'),
        ('xlsx', 'Excel File (.xlsx)'),
        ('xls', 'Excel File (.xls)'),
    ], string='File Type', compute='_compute_file_type', store=True)
    update_existing = fields.Boolean(
        string='Update Existing Records',
        default=True,
        help="Update existing tracker records if found"
    )
    year = fields.Integer(
        string='Year',
        default=lambda self: date.today().year,
        required=True
    )
    import_results = fields.Text(string='Import Results', readonly=True)

    @api.depends('import_filename')
    def _compute_file_type(self):
        for record in self:
            if record.import_filename:
                fname = record.import_filename.lower()
                if fname.endswith('.xlsx'):
                    record.file_type = 'xlsx'
                elif fname.endswith('.xls'):
                    record.file_type = 'xls'
                elif fname.endswith('.csv'):
                    record.file_type = 'csv'
                else:
                    record.file_type = 'csv'
            else:
                record.file_type = 'csv'

    def safe_float(self, value):
        try:
            if value is None:
                return 0.0
            if isinstance(value, str):
                value = value.replace(',', '').strip()
                if value == '':
                    return 0.0
            return float(value)
        except Exception:
            return None

    def action_import_data(self):
        if not self.import_file:
            raise ValidationError(_('Please select a file to import.'))

        try:
            # Parse file
            if self.file_type in ['xlsx', 'xls']:
                if self.file_type == 'xlsx' and not OPENPYXL_AVAILABLE:
                    raise UserError(_('openpyxl is required to process .xlsx files.'))
                if self.file_type == 'xls' and not XLRD_AVAILABLE:
                    raise UserError(_('xlrd is required to process .xls files.'))
                rows, headers = self._parse_excel_file()
            else:
                rows, headers = self._parse_csv_file()

            if not rows:
                raise ValidationError(_('No data found in the file.'))

            imported_count = 0
            updated_count = 0
            errors = []

            for row_num, row in enumerate(rows, start=2):
                try:
                    if not row or (isinstance(row, list) and not any(str(cell).strip() for cell in row)):
                        continue

                    # Map columns to expected order
                    if isinstance(row, dict):
                        employee_id = str(row.get('Employee ID', '')).strip()
                        employee_name = str(row.get('Name', '')).strip()
                        department_name = str(row.get('Department', '')).strip()
                        leave_type_name = str(row.get('Leave Type', '')).strip()
                        year_val = int(row.get('Year', self.year))
                        total_allocated = self.safe_float(row.get('Total Allocation', 0))
                        days_taken = self.safe_float(row.get('Taken Leaves', 0))
                        pending_requests = self.safe_float(row.get('Pending Requests', 0))
                        current_balance = self.safe_float(row.get('Current Balance', 0))
                        carry_forwarded = self.safe_float(row.get('Carry Forwarded', 0))
                        expired_carry = self.safe_float(row.get('Expired Carry', 0))
                    else:
                        if len(row) < 11:
                            errors.append(f"Row {row_num}: Insufficient columns (expected 11, got {len(row)})")
                            continue

                        employee_id = str(row[0] or '').strip()
                        employee_name = str(row[1] or '').strip()
                        department_name = str(row[2] or '').strip()
                        leave_type_name = str(row[3] or '').strip()
                        year_val = int(row[4] or self.year)
                        total_allocated = self.safe_float(row[5])
                        days_taken = self.safe_float(row[6])
                        pending_requests = self.safe_float(row[7])
                        current_balance = self.safe_float(row[8])
                        carry_forwarded = self.safe_float(row[9])
                        expired_carry = self.safe_float(row[10])

                    
                    _logger.info("Processing Row %d: Employee ID='%s', Employee Name='%s'", row_num, employee_id, employee_name)



                    if not employee_id or not leave_type_name:
                        errors.append(f"Row {row_num}: Missing Employee ID or Leave Type")
                        continue

                    employee = self.env['hr.employee'].search([('employee_number', '=', employee_id)], limit=1)
                    if not employee and employee_name:
                        employee = self.env['hr.employee'].search([('name', '=', employee_name)], limit=1)

                    if not employee:
                        errors.append(f"Row {row_num}: Could not find employee {employee_id} ({employee_name})")
                        continue

                    leave_type = self.env['hr.leave.type'].search([('name', '=', leave_type_name)], limit=1)
                    if not leave_type:
                        errors.append(f"Row {row_num}: Leave type '{leave_type_name}' not found")
                        continue

                    department = self.env['hr.department'].search([('name','=',department_name)], limit=1)
                    department_id = department.id if department else False

                    tracker_data = {
                        'employee_id': employee.id,
                        'leave_type_id': leave_type.id,
                        'leave_type_name': leave_type_name,
                        'year': year_val,
                        'total_allocation': total_allocated,
                        'taken_leaves': days_taken,
                        'pending_requests': pending_requests,
                        'current_balance': current_balance,
                        'annual_carry': carry_forwarded,
                        'expired_carry': expired_carry,
                        'employee_name': employee.name,
                        'employee_number': employee.employee_number or '',
                        'department_id': department_id,
                        'name': f"{leave_type_name} {year_val}",
                        # 'is_historical': True,
                    }

                    existing_tracker = self.env['hr.leave.tracker'].search([
                        ('employee_id', '=', employee.id),
                        ('leave_type_id', '=', leave_type.id),
                        ('year', '=', year_val)
                    ], limit=1)

                    if existing_tracker:
                        if self.update_existing:
                            existing_tracker.write(tracker_data)
                            updated_count += 1
                        else:
                            errors.append(f"Row {row_num}: Record exists, skipped")
                    else:
                        self.env['hr.leave.tracker'].create(tracker_data)
                        imported_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            self.env.cr.commit()

            message = f"Import completed for year {self.year}!\n\n"
            message += f"✅ Imported: {imported_count} new records\n"
            message += f"🔄 Updated: {updated_count} existing records\n"
            if errors:
                message += f"\n❌ Errors ({len(errors)}):\n"
                for error in errors[:10]:
                    message += f"• {error}\n"
                if len(errors) > 10:
                    message += f"... and {len(errors) - 10} more errors"
            else:
                message += "\n🎉 All records processed successfully!"

            self.import_results = message
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'hr.leave.import',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }

        except Exception as e:
            raise UserError(_('Error processing file: %s') % str(e))

    def _parse_csv_file(self):
        try:
            file_data = base64.b64decode(self.import_file)
            for enc in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    csv_data = file_data.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise UserError(_('Unable to decode CSV. Use UTF-8.'))

            csv_reader = csv.DictReader(io.StringIO(csv_data))
            rows = list(csv_reader)
            headers = csv_reader.fieldnames or []
            return rows, headers
        except Exception as e:
            raise UserError(_('CSV parsing error: %s') % str(e))

    def _parse_excel_file(self):
        file_data = base64.b64decode(self.import_file)
        try:
            if self.file_type == 'xlsx':
                wb = load_workbook(io.BytesIO(file_data), read_only=True)
                ws = wb.active
                headers = [str(c or '') for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    processed_row = [c if c is not None else '' for c in row]
                    if any(str(c).strip() for c in processed_row):
                        rows.append(processed_row)
                wb.close()
            else:
                wb = xlrd.open_workbook(file_contents=file_data)
                ws = wb.sheet_by_index(0)
                headers = [str(ws.cell_value(0, c) or '') for c in range(ws.ncols)]
                rows = []
                for r in range(1, ws.nrows):
                    row = [ws.cell_value(r, c) or '' for c in range(ws.ncols)]
                    if any(str(c).strip() for c in row):
                        rows.append(row)
            return rows, headers
        except Exception as e:
            raise UserError(_('Excel parsing error: %s') % str(e))

    def action_download_template(self):
        headers = [
            'Employee ID', 'Name', 'Department', 'Leave Type', 'Year',
            'Total Allocation', 'Taken Leaves', 'Pending Requests',
            'Current Balance', 'Carry Forwarded', 'Expired Carry'
        ]
        sample_data = ['EMP001','John Doe','HR','Annual Leave','2025','20','5','2','15','3','0']
        csv_content = ','.join(headers) + '\n' + ','.join(sample_data)
        attachment = self.env['ir.attachment'].create({
            'name': 'hr_leave_template.csv',
            'type': 'binary',
            'datas': base64.b64encode(csv_content.encode('utf-8')),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/csv'
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
